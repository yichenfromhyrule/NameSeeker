# Author @Yichen Wang

import openpyxl
from difflib import SequenceMatcher

def manageSampleExcel(sample_path): #Convert sample.xlsx to a matrix
    m = []
    wb_obj = openpyxl.load_workbook(sample_path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column
    for i in range(1, m_row + 1):
        com = []
        for j in range(1,m_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            if cell_obj.value != None:
                com.append(cell_obj.value)
        m.append(com)
    print(m)
    return m



def getTargetName(test_path, company_name_matrix):
    wb_obj = openpyxl.load_workbook(test_path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        for j in range(0, len(company_name_matrix)):
            for k in range(0, len(company_name_matrix[j])):
                check_name = company_name_matrix[j][k]
                similar_rate = similar(cell_obj.value, check_name)
                if similar_rate == 1:
                    sheet_obj.cell(row=i, column=2).value = company_name_matrix[j][0]

    wb_obj.save("test.xlsx")



def goThroughExcel(l):
    #1. Read the Excel, get the number of rows m_row
    path = "test.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    print("We have %d rows in the Excel."%m_row)
    #2. Use a for loop to go through each row,
    #   calculate the similar rate of each company's name with compang name list

    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        print (cell_obj.value)
        similar_rate_result = []
        for j in range(0, len(l)):
            similar_rate_result.append(similar(cell_obj.value, l[j]))
        print (cell_obj, similar_rate_result)
        print (similar_rate_result.index(max(similar_rate_result)))
        target_name = l[similar_rate_result.index(max(similar_rate_result))]
        sheet_obj.cell(row = i, column = 2).value = target_name
        sheet_obj.cell(row=i, column=3).value = max(similar_rate_result)
    wb_obj.save("test.xlsx")

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    sampleExcelPath = "sample.xlsx"
    testExcelPath = "test.xlsx"
    company_name_matrix = manageSampleExcel(sampleExcelPath)
    getTargetName(testExcelPath, company_name_matrix)

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
